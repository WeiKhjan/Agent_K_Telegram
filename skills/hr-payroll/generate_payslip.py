#!/usr/bin/env python3
"""
AiTraining2U PLT — Payslip Generator
Generates one Excel workbook per employee with a professional payslip layout.
Reusable template: add employees to the PAYSLIP_DATA list.

Usage:
  python3 generate_payslip.py
  python3 generate_payslip.py --month 2026-03 --out ~/payslips/

Required: /Users/aitraining2u/.local/share/office-venv/bin/python
"""

import argparse
import os
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not found. Use office-venv python.")
    sys.exit(1)

# ── Company Info ─────────────────────────────────────────────────────────────
COMPANY = {
    "name": "AiTraining2U PLT",
    "reg":  "202504002669 (LLP0051040-LGN)",
    "addr": "D-10-5, Sky Condominium, Persiaran Puchong Jaya Selatan, 47100 Selangor",
    "email": "atlas.aitraining2u@gmail.com",
}

# ── Statutory Rates ──────────────────────────────────────────────────────────
EPF_EMP_RATE   = 0.11   # employee 11%
EPF_ER_RATE    = 0.15   # employer 15% (AiTraining2U company policy, per employment letter)
SOCSO_EMP_RATE = 0.005  # employee 0.5%
SOCSO_ER_RATE  = 0.0175 # employer 1.75%
EIS_EMP_RATE   = 0.004  # employee 0.4%
EIS_ER_RATE    = 0.004  # employer 0.4%


# ── Styles ───────────────────────────────────────────────────────────────────
NAVY      = "1F3864"
NAVY_LIGHT= "D6E4F7"
TEAL      = "1ABC9C"
GREY_BG   = "F5F5F5"
WHITE     = "FFFFFF"
ACCENT    = "2E75B6"

def side(style="thin", color="AAAAAA"):
    return Side(style=style, color=color)

THIN_BORDER = Border(left=side(), right=side(), top=side(), bottom=side())
MED_BORDER  = Border(left=side("medium","1F3864"), right=side("medium","1F3864"),
                     top=side("medium","1F3864"), bottom=side("medium","1F3864"))

def hdr_font(bold=True, size=10, color=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def body_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def right():
    return Alignment(horizontal="right", vertical="center")

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

MYR_FMT = '#,##0.00'
PCT_FMT = '0.0%'


# ── Helpers ───────────────────────────────────────────────────────────────────
def merge_write(ws, cell_range, value, font=None, fill_=None, align=None, border=None, num_fmt=None):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = value
    if font:   c.font      = font
    if fill_:  c.fill      = fill_
    if align:  c.alignment = align
    if border: c.border    = border
    if num_fmt:c.number_format = num_fmt
    return c

def write(ws, row, col, value, font=None, fill_=None, align=None, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill_:  c.fill      = fill_
    if align:  c.alignment = align
    if border: c.border    = border
    if num_fmt:c.number_format = num_fmt
    return c

def section_header(ws, row, col_start, col_end, text):
    rng = f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}"
    ws.merge_cells(rng)
    c = ws[f"{get_column_letter(col_start)}{row}"]
    c.value     = text
    c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill      = fill(ACCENT)
    c.alignment = left()
    c.border    = THIN_BORDER
    ws.row_dimensions[row].height = 18

def row_pair(ws, row, label, amount_formula_or_val, col_label=2, col_val=5,
             bold=False, total=False, bg=None):
    """Write a label-value pair row."""
    bg_hex = bg or (GREY_BG if row % 2 == 0 else WHITE)
    f = fill(bg_hex)
    bf = body_font(bold=bold or total)

    ws.merge_cells(f"B{row}:D{row}")
    lc = ws[f"B{row}"]
    lc.value     = label
    lc.font      = bf
    lc.fill      = f
    lc.alignment = left()
    lc.border    = THIN_BORDER

    vc = ws[f"E{row}"]
    vc.value     = amount_formula_or_val
    vc.font      = bf
    vc.fill      = f
    vc.alignment = right()
    vc.border    = THIN_BORDER
    vc.number_format = MYR_FMT

    ws.row_dimensions[row].height = 16
    return vc


# ── Payslip builder ───────────────────────────────────────────────────────────
def build_payslip(ws, emp: dict, month_str: str):
    """
    emp dict keys:
      name, emp_id, position, basic, ot, allowances (list of (label, amount)),
      pcb, tax_cat, dependents
    month_str: "2026-03"
    """
    dt = datetime.strptime(month_str + "-01", "%Y-%m-%d")
    month_label = dt.strftime("%B %Y")   # "March 2026"

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 2

    # ── Header block ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8

    merge_write(ws, "B2:E2", COMPANY["name"],
                font=Font(name="Calibri", bold=True, size=16, color=NAVY),
                fill_=fill(WHITE), align=left())
    merge_write(ws, "B3:E3", COMPANY["reg"],
                font=body_font(size=9, color="555555"), align=left())
    merge_write(ws, "B4:E4", COMPANY["addr"],
                font=body_font(size=9, color="555555"), align=left(wrap=True))
    ws.row_dimensions[4].height = 24

    # Divider
    for col in range(2, 6):
        c = ws.cell(row=5, column=col)
        c.border = Border(bottom=Side(style="thick", color=NAVY))
    ws.row_dimensions[5].height = 6

    # Payslip title + month
    merge_write(ws, "B6:C6", "PAYSLIP",
                font=Font(name="Calibri", bold=True, size=14, color=NAVY),
                align=left())
    merge_write(ws, "D6:E6", month_label,
                font=body_font(bold=True, size=11, color=ACCENT),
                align=right())
    ws.row_dimensions[6].height = 22

    ws.row_dimensions[7].height = 6  # spacer

    # ── Employee Info ─────────────────────────────────────────────────────────
    section_header(ws, 8, 2, 5, "  EMPLOYEE INFORMATION")

    info_rows = [
        ("Employee Name",  emp["name"]),
        ("Employee ID",    emp["emp_id"]),
        ("Position",       emp["position"]),
        ("Employment Type","Full-Time (Permanent)"),
        ("Pay Period",     month_label),
    ]
    for i, (lbl, val) in enumerate(info_rows):
        r = 9 + i
        bg = GREY_BG if i % 2 == 0 else WHITE
        ws.merge_cells(f"B{r}:C{r}")
        lc = ws[f"B{r}"]
        lc.value     = lbl
        lc.font      = body_font(bold=True)
        lc.fill      = fill(bg)
        lc.alignment = left()
        lc.border    = THIN_BORDER

        ws.merge_cells(f"D{r}:E{r}")
        vc = ws[f"D{r}"]
        vc.value     = val
        vc.font      = body_font()
        vc.fill      = fill(bg)
        vc.alignment = left()
        vc.border    = THIN_BORDER
        ws.row_dimensions[r].height = 16

    r = 9 + len(info_rows)  # next row

    ws.row_dimensions[r].height = 8  # spacer
    r += 1

    # ── Earnings ──────────────────────────────────────────────────────────────
    section_header(ws, r, 2, 5, "  EARNINGS")
    r += 1
    earn_start = r

    basic_row = r
    row_pair(ws, r, "Basic Salary", emp["basic"]); r += 1

    ot_row = None
    if emp.get("ot", 0):
        ot_row = r
        row_pair(ws, r, "Overtime Pay", emp["ot"]); r += 1

    allow_rows = []
    for (lbl, amt) in emp.get("allowances", []):
        allow_rows.append(r)
        row_pair(ws, r, lbl, amt); r += 1

    # Gross total
    gross_row = r
    gross_cells = [f"E{basic_row}"]
    if ot_row: gross_cells.append(f"E{ot_row}")
    for ar in allow_rows: gross_cells.append(f"E{ar}")
    gross_formula = "=" + "+".join(gross_cells)

    row_pair(ws, r, "GROSS EARNINGS", gross_formula, bold=True, bg=NAVY_LIGHT)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    r += 1

    gross_ref = f"E{gross_row}"

    ws.row_dimensions[r].height = 8; r += 1

    # ── Deductions ────────────────────────────────────────────────────────────
    section_header(ws, r, 2, 5, "  DEDUCTIONS (Employee)")
    r += 1

    epf_row = r
    epf_amt = round(emp["basic"] * EPF_EMP_RATE)  # EPF on basic only (standard)
    row_pair(ws, r, f"EPF / KWSP (11%)", epf_amt); r += 1

    # SOCSO on gross (capped at RM 5,000)
    socso_row = r
    insurable = min(emp["basic"] + emp.get("ot", 0), 5000)
    socso_amt = round(insurable * SOCSO_EMP_RATE, 2)
    row_pair(ws, r, "SOCSO / PERKESO (0.5%)", socso_amt); r += 1

    # EIS on gross (capped at RM 5,000)
    eis_row = r
    eis_insurable = min(emp["basic"] + emp.get("ot", 0), 5000)
    eis_amt = round(eis_insurable * EIS_EMP_RATE, 2)
    row_pair(ws, r, "EIS / SIP (0.4%)", eis_amt); r += 1

    # PCB / MTD
    pcb_row = r
    pcb_val = emp.get("pcb", 0)
    pcb_label = "PCB / MTD (Income Tax)"
    if pcb_val == 0:
        pcb_label += " *"
    row_pair(ws, r, pcb_label, pcb_val); r += 1

    # Total deductions
    total_ded_row = r
    ded_formula = f"=E{epf_row}+E{socso_row}+E{eis_row}+E{pcb_row}"
    row_pair(ws, r, "TOTAL DEDUCTIONS", ded_formula, bold=True, bg=NAVY_LIGHT)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    r += 1

    ws.row_dimensions[r].height = 8; r += 1

    # ── Net Pay ───────────────────────────────────────────────────────────────
    section_header(ws, r, 2, 5, "  NET PAY")
    r += 1

    net_row = r
    net_formula = f"={gross_ref}-E{total_ded_row}"
    c = ws.cell(row=r, column=2)
    ws.merge_cells(f"B{r}:D{r}")
    c.value     = "NET SALARY (Take-Home Pay)"
    c.font      = Font(name="Calibri", bold=True, size=12, color=WHITE)
    c.fill      = fill(NAVY)
    c.alignment = left()
    c.border    = MED_BORDER

    vc = ws.cell(row=r, column=5)
    vc.value         = net_formula
    vc.font          = Font(name="Calibri", bold=True, size=12, color=WHITE)
    vc.fill          = fill(NAVY)
    vc.alignment     = right()
    vc.border        = MED_BORDER
    vc.number_format = MYR_FMT
    ws.row_dimensions[r].height = 24
    r += 1

    ws.row_dimensions[r].height = 8; r += 1

    # ── Employer Contributions (info only) ────────────────────────────────────
    section_header(ws, r, 2, 5, "  EMPLOYER CONTRIBUTIONS (For Reference)")
    r += 1

    epf_er = round(emp["basic"] * EPF_ER_RATE)
    socso_er = round(insurable * SOCSO_ER_RATE, 2)
    eis_er   = round(eis_insurable * EIS_ER_RATE, 2)

    er_rows = [
        (f"EPF / KWSP Employer (13%)", epf_er),
        ("SOCSO / PERKESO Employer (1.75%)", socso_er),
        ("EIS / SIP Employer (0.4%)", eis_er),
    ]
    for i, (lbl, amt) in enumerate(er_rows):
        row_pair(ws, r, lbl, amt); r += 1

    er_total_row = r
    row_pair(ws, r, "Total Employer Contributions",
             f"=E{r-3}+E{r-2}+E{r-1}", bold=True, bg=NAVY_LIGHT)
    r += 1

    ws.row_dimensions[r].height = 8; r += 1

    # ── PCB note ──────────────────────────────────────────────────────────────
    if pcb_val == 0:
        ws.merge_cells(f"B{r}:E{r}")
        nc = ws[f"B{r}"]
        nc.value     = "* PCB/MTD: Pending — tax category & dependents not yet collected from employee."
        nc.font      = Font(name="Calibri", italic=True, size=9, color="AA4400")
        nc.alignment = left()
        ws.row_dimensions[r].height = 14
        r += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    r += 2
    merge_write(ws, f"B{r}:E{r}",
                "This payslip is computer generated. No signature required if printed.",
                font=body_font(size=8, color="888888"), align=center())

    # Print settings
    ws.print_area = f"A1:F{r+1}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left  = 0.5
    ws.page_margins.right = 0.5


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="AiTraining2U Payslip Generator")
    parser.add_argument("--month", default="2026-03", help="YYYY-MM format")
    parser.add_argument("--out",   default=os.path.expanduser("~/"), help="Output directory")
    args = parser.parse_args()

    out_dir = Path(args.out).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Payslip data: edit this section each month ────────────────────────────
    PAYSLIP_DATA = [
        {
            "name":       "Farah Safiyyah Binti Md Zahidan",
            "emp_id":     "ATU-EMP-2026-0005",
            "position":   "AI Graduate Trainee",
            "basic":      3800.00,
            "ot":         300.00,
            "allowances": [],       # e.g. [("Training Allowance", 400)]
            "pcb":        0,        # 0 = pending; update when collected
        },
        {
            "name":       "Danial Syafiq Bin Amir Hashim",
            "emp_id":     "ATU-EMP-2026-0006",
            "position":   "AI Graduate Trainee",
            "basic":      3800.00,
            "ot":         300.00,
            "allowances": [("Training Allowance", 400.00)],
            "pcb":        0,
        },
        {
            "name":       "Muhammad Fareed 'Aidil Bin Rozaidi",
            "emp_id":     "ATU-EMP-2026-0003",
            "position":   "AI Graduate Trainee",
            "basic":      3800.00,
            "ot":         0,
            "allowances": [],
            "pcb":        0,
        },
    ]
    # ─────────────────────────────────────────────────────────────────────────

    generated = []
    month_label = datetime.strptime(args.month + "-01", "%Y-%m-%d").strftime("%B_%Y")

    for emp in PAYSLIP_DATA:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payslip"

        build_payslip(ws, emp, args.month)

        short_name = emp["name"].split()[0]  # first name
        filename = f"Payslip_{args.month}_{short_name}_{emp['emp_id']}.xlsx"
        out_path = out_dir / filename
        wb.save(out_path)
        generated.append(str(out_path))
        print(f"  ✓ {filename}")

    print(f"\nGenerated {len(generated)} payslip(s) in {out_dir}")
    return generated


if __name__ == "__main__":
    main()
